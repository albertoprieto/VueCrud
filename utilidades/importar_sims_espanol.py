"""Carga el Excel "SIM ESPAÑOL" a la tabla consultas_sim de MI base de datos.

El Excel es la verdad: para cada fila se busca el registro por ICCID y se
SETEAN los campos listados abajo con lo que diga el Excel. NO toca SIMPRO ni
ninguna columna ligada a SIMPRO (sim_state, salud, network_imei, ...): solo
escribe columnas planas de consultas_sim vía UPDATE directo.

Columnas que sobreescribe (llave = iccid):
    activation_date, tipo (SERVICIO), deaccount (USUARIO), account_name (CLIENTE),
    plataforma, imei, device_mobile (SIM ESPAÑOL), vigencia_sim,
    tecnico (TECNICO), num_cliente (NUM. CLIENTE), comentarios (COMENTARIOS)

vigencia_sim = activation_date + 365 días (misma fórmula del Excel: =A+365).
"Días restantes" NO se guarda: la calcula el front (vigencia - hoy).

ICCID repetido en el Excel = historial de renovaciones -> se importa solo la
fila con activation_date más reciente (la DB tiene 1 fila por ICCID).
ICCID con <18 dígitos = corrupto en el Excel (guardado como número) -> se ignora.

Por defecto una celda vacía del Excel NO pisa el valor de la DB. Con --vaciar
se respeta el Excel al pie de la letra (deja el campo vacío).

--insertar : los ICCID del Excel que no existen en la DB se INSERTAN como
registros nuevos (los que no traen ni IMEI ni SIM español se omiten). El
revert también borra los insertados (tabla <respaldo>_nuevos).

Uso:
    python importar_sims_espanol.py sims_espanol_actualizado.xlsx           # dry-run
    python importar_sims_espanol.py sims_espanol_actualizado.xlsx --aplicar # escribe + crea respaldo
    python importar_sims_espanol.py --revert consultas_sim_respaldo_20260831_120000

Al aplicar se crea una tabla consultas_sim_respaldo_<timestamp> con copia
completa de las filas tocadas. --revert restaura desde ahí (solo las columnas
de arriba) y deja la tabla de respaldo por si acaso.
"""
import json
import sys
from datetime import datetime, timedelta
from pathlib import Path

import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook

import os  # noqa: E402

_aqui = Path(__file__).resolve().parent
for _cand in (_aqui / ".env", _aqui.parent / ".env", Path.cwd() / ".env"):
    if _cand.is_file():
        load_dotenv(_cand)
        break
else:
    load_dotenv()

# Columnas que este script controla. Nada fuera de esta lista se toca.
COLS = [
    "activation_date", "tipo", "deaccount", "account_name", "plataforma",
    "imei", "device_mobile", "vigencia_sim", "tecnico", "num_cliente", "comentarios",
]

MAP_TIPO = {
    "ACTIVACION": "activacion", "ACTIVACIÓN": "activacion", "NO ACTIVADO": "activacion",
    "RENOVACION": "renovacion", "RENOVACIÓN": "renovacion", "RENOVADO": "renovacion",
    "REINSTALACION": "reutilizado", "REINSTALACIÓN": "reutilizado",
    "REUTILIZADO": "reutilizado",
    "CANCELADO": "cancelado", "DESINSTALADO": "desinstalado",
}
VACIO = {"", "SIN DATOS", "SIN DATO", "-", "N/A", "NA", "NONE", "NULL"}


def db():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
        database=os.getenv("DB_NAME"),
    )


def digits(v):
    if isinstance(v, float) and v.is_integer():
        v = int(v)  # openpyxl entrega números como float: 4477...91.0 -> 4477...91
    return "".join(ch for ch in str(v or "") if ch.isdigit())


def limpio(v):
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v if v is not None else "").strip()
    return "" if s.upper() in VACIO else s


def norm_plat(v):
    p = limpio(v).upper()
    if p in {"TRACKSOLID", "TRACK", "TRACKSOLIDPRO"}:
        return "TRACKSOLID"
    if p in {"IOP", "IOPGPS", "IOPSGPS", "IOP GPS", "WANWAY"}:
        return "IOP"
    return p


def a_fecha_iso(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = limpio(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return ""


def mas_365(fecha_iso):
    if not fecha_iso:
        return ""
    return (datetime.strptime(fecha_iso, "%Y-%m-%d") + timedelta(days=365)).strftime("%Y-%m-%d")


def leer_excel(ruta):
    ws = load_workbook(ruta, data_only=True).active
    filas = list(ws.iter_rows(values_only=True))
    hdr_idx = next(
        i for i, r in enumerate(filas)
        if r and any(str(c).strip().upper() == "ICCID" for c in r if c)
        and any("SERVICIO" in str(c).upper() for c in r if c)
    )
    hdr = [str(c or "").strip().upper().replace("ESPA?OL", "ESPAÑOL") for c in filas[hdr_idx]]

    def col(nombre):
        return hdr.index(nombre)

    ix = {
        "fecha": col("FECHA. ACT"), "servicio": col("SERVICIO"), "usuario": col("USUARIO"),
        "cliente": col("CLIENTE"), "plataforma": col("PLATAFORMA"), "imei": col("IMEI"),
        "iccid": col("ICCID"), "sim": col("SIM ESPAÑOL"), "tecnico": col("TECNICO"),
        "comentarios": col("COMENTARIOS"), "num_cliente": col("NUM. CLIENTE"),
    }
    out = []
    for r in filas[hdr_idx + 1:]:
        if not r or not any(r):
            continue
        iccid = digits(r[ix["iccid"]])
        if not iccid:
            continue
        act = a_fecha_iso(r[ix["fecha"]])
        imei = digits(r[ix["imei"]])
        if len(imei) < 14:  # celda con basura ("X", "SIN DATOS", etc.)
            imei = ""
        out.append({
            "iccid": iccid,
            "activation_date": act,
            "tipo": MAP_TIPO.get(limpio(r[ix["servicio"]]).upper(), "activacion"),
            "deaccount": limpio(r[ix["usuario"]]),
            "account_name": limpio(r[ix["cliente"]]),
            "plataforma": norm_plat(r[ix["plataforma"]]),
            "imei": imei,
            "device_mobile": digits(r[ix["sim"]]),
            "vigencia_sim": mas_365(act),
            "tecnico": limpio(r[ix["tecnico"]]),
            "num_cliente": limpio(r[ix["num_cliente"]]),
            "comentarios": limpio(r[ix["comentarios"]]),
        })

    # ICCID corrupto: Excel lo guardó como número y perdió dígitos.
    corruptos = sorted({e["iccid"] for e in out if len(e["iccid"]) < 18})
    out = [e for e in out if len(e["iccid"]) >= 18]

    # ICCID repetido = historial de renovaciones. La DB tiene 1 fila por ICCID:
    # nos quedamos con la activación/renovación MÁS RECIENTE (activation_date máx).
    por_iccid = {}
    for e in out:
        prev = por_iccid.get(e["iccid"])
        if prev is None or e["activation_date"] > prev["activation_date"]:
            por_iccid[e["iccid"]] = e
    historial = len(out) - len(por_iccid)
    return list(por_iccid.values()), {"corruptos": corruptos, "filas_historial": historial}


def norm_actual(row):
    """Valores actuales en DB, normalizados igual que el Excel para comparar."""
    return {c: (a_fecha_iso(row[c]) if c in ("activation_date", "vigencia_sim")
               else str(row[c] or "").strip()) for c in COLS}


def asegurar_columnas(conn):
    """Crea tecnico / num_cliente / comentarios si el backend aún no migró."""
    cur = conn.cursor()
    for ddl in (
        "ALTER TABLE consultas_sim ADD COLUMN tecnico VARCHAR(120) NULL",
        "ALTER TABLE consultas_sim ADD COLUMN num_cliente VARCHAR(60) NULL",
        "ALTER TABLE consultas_sim ADD COLUMN comentarios TEXT NULL",
    ):
        try:
            cur.execute(ddl)
        except mysql.connector.Error:
            pass
    conn.commit()
    cur.close()


def aplicar(ruta, escribir, vaciar, insertar):
    conn = db()
    asegurar_columnas(conn)
    cur = conn.cursor(dictionary=True)
    excel, avisos = leer_excel(ruta)
    print(f"SIMs (ICCID únicos) en el Excel: {len(excel)}")
    print(f"  filas de historial de renovación descartadas: {avisos['filas_historial']}")
    print(f"  ICCID corruptos en el Excel (guardados como número): {len(avisos['corruptos'])}")
    if avisos["corruptos"]:
        print("   ", ", ".join(avisos["corruptos"]))
    print(f"  modo campos vacíos: {'SETEAR vacío (Excel manda)' if vaciar else 'conservar valor de DB'}")

    cambios, sin_match, nuevos, sin_cambio, choques = [], [], [], 0, []
    for e in excel:
        cur.execute(
            "SELECT id, " + ", ".join(COLS) + " FROM consultas_sim "
            "WHERE REPLACE(iccid,' ','') = %s LIMIT 1",
            (e["iccid"],),
        )
        row = cur.fetchone()
        if not row:
            sin_match.append(e["iccid"])
            if insertar:
                if e["imei"] or e["device_mobile"]:
                    nuevos.append(e)
                else:
                    choques.append({"iccid": e["iccid"], "error": "sin imei ni sim, no se inserta"})
            continue
        actual = norm_actual(row)
        nuevo = {c: (e[c] if (e[c] or vaciar) else actual[c]) for c in COLS}
        if actual == nuevo:
            sin_cambio += 1
            continue
        cambios.append((row["id"], nuevo, actual))

    print(f"  a actualizar: {len(cambios)}")
    print(f"  sin cambios:  {sin_cambio}")
    print(f"  sin match en DB: {len(sin_match)}"
          + (f"  ->  a INSERTAR: {len(nuevos)}" if insertar else "  (usa --insertar para crearlos)"))
    if sin_match:
        print("  ICCID sin match:", ", ".join(sin_match[:20]), "..." if len(sin_match) > 20 else "")

    if not escribir:
        print("\nDRY-RUN. Nada escrito. Corre con --aplicar para ejecutar.")
        for cid, nuevo, viejo in cambios[:10]:
            difs = {c: (viejo[c], nuevo[c]) for c in COLS if str(viejo[c] or "").strip() != nuevo[c]}
            print(f"  id={cid} {difs}")
        for e in nuevos[:10]:
            print(f"  NUEVO iccid={e['iccid']} {[e[c] for c in COLS]}")
        return

    respaldo = f"consultas_sim_respaldo_{datetime.now():%Y%m%d_%H%M%S}"
    ids = [str(cid) for cid, _, _ in cambios]
    w = conn.cursor()
    w.execute(f"CREATE TABLE {respaldo} LIKE consultas_sim")
    if ids:
        w.execute(
            f"INSERT INTO {respaldo} SELECT * FROM consultas_sim WHERE id IN ({','.join(ids)})"
        )
    sets = ", ".join(f"{c}=%s" for c in COLS)
    ok = 0
    for cid, nuevo, _ in cambios:
        try:
            w.execute(f"UPDATE consultas_sim SET {sets} WHERE id=%s",
                      [nuevo[c] for c in COLS] + [cid])
            ok += 1
        except mysql.connector.Error as err:
            choques.append({"id": cid, "error": str(err)})

    insertados = 0
    cols_ins = ["iccid"] + COLS
    ph = ", ".join(["%s"] * len(cols_ins))
    w.execute(f"CREATE TABLE {respaldo}_nuevos (id INT)")
    for e in nuevos:
        try:
            w.execute(f"INSERT INTO consultas_sim ({', '.join(cols_ins)}) VALUES ({ph})",
                      [e[c] for c in cols_ins])
            w.execute(f"INSERT INTO {respaldo}_nuevos (id) VALUES (%s)", (w.lastrowid,))
            insertados += 1
        except mysql.connector.Error as err:
            choques.append({"iccid": e["iccid"], "error": str(err)})
    conn.commit()
    print(json.dumps({
        "respaldo": respaldo,
        "actualizados": ok,
        "insertados": insertados,
        "errores": choques,
        "revertir_con": f"python importar_sims_espanol.py --revert {respaldo}",
    }, ensure_ascii=False, indent=2))


def revertir(tabla):
    conn = db()
    cur = conn.cursor(dictionary=True)
    cur.execute(f"SELECT id, {', '.join(COLS)} FROM {tabla}")
    filas = cur.fetchall()
    w = conn.cursor()
    sets = ", ".join(f"{c}=%s" for c in COLS)
    for row in filas:
        w.execute(f"UPDATE consultas_sim SET {sets} WHERE id=%s",
                  [row[c] for c in COLS] + [row["id"]])

    borrados = 0
    try:
        cur.execute(f"SELECT id FROM {tabla}_nuevos")
        ids = [str(r["id"]) for r in cur.fetchall()]
        if ids:
            w.execute(f"DELETE FROM consultas_sim WHERE id IN ({','.join(ids)})")
            borrados = w.rowcount
    except mysql.connector.Error:
        pass  # no hubo inserts en esa corrida
    conn.commit()
    print(f"Revertidas {len(filas)} filas y borradas {borrados} insertadas desde {tabla}. "
          "Las tablas de respaldo se conservan.")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        sys.exit(1)
    if args[0] == "--revert":
        revertir(args[1])
    else:
        aplicar(args[0], escribir="--aplicar" in args, vaciar="--vaciar" in args,
                insertar="--insertar" in args)
